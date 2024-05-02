# Translated from Company Scrapping_yfinance_Validation.ipynb
# Third-party Libraries
import tagui as t
import yfinance as yf
from datetime import datetime
# import datetime
from openpyxl import Workbook, load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from nltk.sentiment.vader import SentimentIntensityAnalyzer
import dotenv

# Native Libraries
from itertools import combinations
import re
import requests

dotenv.load_dotenv()
# dotenv.load_dotenv(r'RAG/.env')

################ 1. SCRAPPING VALIDATION ################
def compare_attributes(tagui_value, yfinance_value, attribute_name, diff_count):
    """
    Desc: Function to compare attributes and count differences

    Input
    ------
    tagui_value: info scrapped by TagUI
    yfinance_value: info taken from yahoo finance's API
    attribute_name: the name of info
    diff_count: Number of differences in info between TagUI and yahoo finance (accumulative)

    Return
    ------
    diff_count: the resulting number of diff
    """
    # Check if the tagui_value is a string
    if isinstance(tagui_value, str):
        # Remove commas from the string
        tagui_value = tagui_value.replace(',', '')

        if tagui_value.endswith('B'):
            tagui_value = float(tagui_value[:-1]) * 1e9  # Convert billions to numeric value
            yfinance_value = round(yfinance_value / 1000000) * 1000000
        elif tagui_value.endswith('M'):
            tagui_value = float(tagui_value[:-1]) * 1e6  # Convert millions to numeric value
            yfinance_value = round(yfinance_value / 1000) * 1000
        else:
            tagui_value = float(tagui_value)  # Return numeric value as is if no suffix is present
   
    tagui_value_rounded = round(float(tagui_value), 2)
    yfinance_value_rounded = round(float(yfinance_value), 2)
    if tagui_value_rounded != yfinance_value_rounded:  # Compare attribute values
        print(f"Difference found for {attribute_name}: TagUI: {tagui_value_rounded}; yFinance: {yfinance_value_rounded}")
        diff_count += 1  # Increment difference count
    # else:
    #     print(f"Values aligned for {attribute_name}: TagUI: {tagui_value}; yFinance: {yfinance_value}")
    
    return diff_count


def tagui_yfinance_validation(company):
    """
    Desc: To confirm if information between info scrapped from TagUI and yahoo finance is the same, \
    and return the number of differences in info between the 2

    Input
    ------
    company (string): Name of company to look at

    Output
    ------

    """
    try:
        # Scraping by TagUI
        t.close()
        t.init(visual_automation = True) # visual automation if keyboard automation required in subsequent code
        t.url('https://sg.finance.yahoo.com/') # go to google finance website
        t.click('//*[@id="yfin-usr-qry"]') # click on search bar.
        t.type('//*[@id="yfin-usr-qry"]',  company+'[enter]') # search for company.
        t.click('//*[contains(@data-id,"result-quotes-0")]') # click the top return company. if above enter does not work
        ticker = t.read('//*[@id="quote-header-info"]//h1').split("(")[1].split(")")[0] # note this ticker symbol includes the full symbol in paranthesis on yahoo finance webpage so that it feeds correctly to yfinance py package
        live_price = float(t.read('//*[@id="quote-header-info"]//*[contains(@data-field,"regularMarketPrice")]')) # reads live stock exchange price
        # stock_market = t.read('//*[@id="quote-header-info"]//div[contains(@class,"Fz(12px)")]/span').split("-")[0].split(" ")[0]
        # Split the string based on the word "Currency"
        split_string = t.read('//*[@id="quote-header-info"]//div[contains(@class,"Fz(12px)")]/span').split("Currency in ")
        # currency = split_string[1].strip()  # Remove any leading or trailing spaces
        prev_close = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[1]/td[2]')
        open_price = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[2]/td[2]')
        trade_daily_volume = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[7]/td[2]')
        day_price_range = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[5]/td[2]')
        day_low = day_price_range.split(" - ")[0]
        day_high = day_price_range.split(" - ")[1]
        market_cap = t.read('//*[contains(@data-test,"right-summary-table")]//table/tbody/tr[1]/td[2]')
        pe_ratio_ttm = t.read('//*[contains(@data-test,"right-summary-table")]//table/tbody/tr[3]/td[2]')
        eps_ttm = t.read('//*[contains(@data-test,"right-summary-table")]//table/tbody/tr[4]/td[2]')

        # Scraping from yfinance. Input ticker using above TagUI
        stock = yf.Ticker(ticker)
        summary_data = stock.info
        yf_live_price = stock.history(period="1d")['Close'].iloc[-1]
        # yf_stock_market = summary_data.get("exchange")
        # yf_currency = summary_data.get("currency")
        yf_prev_close = summary_data.get("previousClose")
        yf_open_price = summary_data.get("open")
        yf_trade_daily_volume = summary_data.get("volume")
        yf_day_low = summary_data.get("dayLow")
        yf_day_high = summary_data.get("dayHigh")
        yf_market_cap = summary_data.get("marketCap")
        yf_pe_ratio_ttm = summary_data.get("trailingPE")
        yf_eps_ttm = summary_data.get("trailingEps")
        # Calculate day_price_range
        # yf_day_price_range = f"{day_low} - {day_high}"

        
        diff_count = 0  # Initialize difference count

        # Compare attributes and count differences
        diff_count = compare_attributes(live_price, yf_live_price, "Live Price", diff_count)
        # diff_count = compare_attributes(stock_market, yf_stock_market, "Stock Market", diff_count)
        # diff_count = compare_attributes(currency, yf_currency, "Currency", diff_count)
        diff_count = compare_attributes(prev_close, yf_prev_close, "Previous Close", diff_count)
        diff_count = compare_attributes(open_price, yf_open_price, "Open Price", diff_count)
        diff_count = compare_attributes(trade_daily_volume, yf_trade_daily_volume, "Trade Daily Volume", diff_count)
        # diff_count = compare_attributes(day_price_range, yf_day_price_range, "Day Price Range", diff_count)
        diff_count = compare_attributes(day_low, yf_day_low, "Day Low", diff_count)
        diff_count = compare_attributes(day_high, yf_day_high, "Day High", diff_count)
        diff_count = compare_attributes(market_cap, yf_market_cap, "Market Cap", diff_count)
        diff_count = compare_attributes(pe_ratio_ttm, yf_pe_ratio_ttm, "PE Ratio TTM", diff_count)
        diff_count = compare_attributes(eps_ttm, yf_eps_ttm, "EPS TTM", diff_count)

        # Output the result
        if diff_count == 0:
            print("No differences found. Values are aligned.")
        else:
            print(f"Total {diff_count} differences found.")
            
        return 
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        return None



################ 2. Chart Snapshot ################
def chart_snapshot(ticker, chart_duration):
    """
    Desc: snap and save stock price chart

    Input
    ------
    ticker: ???
    chart_duration: 
    """

    for duration in chart_duration:
        t.click('//button[contains(text(), "' + duration + '")]') # click duration
        formatted_time = datetime.now().strftime("%Y%m%d_%H%M")
        t.snap('//canvas[contains(@aria-label,"trendArea chart")]', ticker + '/' + duration + '_chart_' + formatted_time + '.png')

    return 



################ 3. Extract company info with Yahoo Finance ################
def extract_company_info_yfinance(company):
    """
    Desc: Extract Company Info from Yahoo Finance

    Input
    ------
    company (string): name of company to look at
    chart_duration (list): duration e.g. ['1d', '1y']
    """
    try:
        # Step 1: to access yahoo finance webpage and key in company of interest
        t.init(visual_automation = True) # visual automation if keyboard automation required in subsequent code
        t.url('https://sg.finance.yahoo.com/') # go to google finance website
        t.click('//*[@id="yfin-usr-qry"]') # click on search bar.
        t.type('//*[@id="yfin-usr-qry"]',  company+'[enter]') # search for company.
        # t.click('//*[contains(@data-id,"result-quotes-0")]') # click the top return company. if above enter does not work
        t.wait(2)
        
        # Step 2: scrape from main summary page
        live_price = float(t.read('//*[@id="quote-header-info"]//*[contains(@data-field,"regularMarketPrice")]')) # reads live stock exchange price
        ticker = t.read('//*[@id="quote-header-info"]//h1').split("(")[1].split(".")[0].split(")")[0]
        ticker_yf = t.read('//*[@id="quote-header-info"]//h1').split("(")[1].split(")")[0]
        stock_market = t.read('//*[@id="quote-header-info"]//div[contains(@class,"Fz(12px)")]/span').split("-")[0].split(" ")[0]
        # Split the string based on the word "Currency"
        split_string = t.read('//*[@id="quote-header-info"]//div[contains(@class,"Fz(12px)")]/span').split("Currency in ")
        currency = split_string[1].strip()  # Remove any leading or trailing spaces
        prev_close = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[1]/td[2]')
        open_price = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[2]/td[2]')
        trade_daily_volume = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[7]/td[2]')
        day_price_range = t.read('//*[contains(@data-test,"left-summary-table")]//table/tbody/tr[5]/td[2]')
        day_low = day_price_range.split(" - ")[0]
        day_high = day_price_range.split(" - ")[1]
        market_cap = t.read('//*[contains(@data-test,"right-summary-table")]//table/tbody/tr[1]/td[2]')
        pe_ratio_ttm = t.read('//*[contains(@data-test,"right-summary-table")]//table/tbody/tr[3]/td[2]')
        eps_ttm = t.read('//*[contains(@data-test,"right-summary-table")]//table/tbody/tr[4]/td[2]')
        # chart_snapshot(ticker, chart_duration)
        
        # Step 3: scrape from Company Profile page
        t.click('//*[contains(@data-test,"COMPANY_PROFILE")]') # click the company profile tab. 
        t.wait(3)
        Company_Full = t.read('//div[@id="Main"]//div/h3') # read full name of company. act as confirmation to user's request as the search name used may not be complete
        industry = t.read('//*[@id="Col1-0-Profile-Proxy"]//span[contains(@class,"Fw(600)")][2]') # read industry of company
        employee_count = t.read('//*[@id="Col1-0-Profile-Proxy"]//span[contains(@class,"Fw(600)")][3]') # read number of full-time employees of company.
        about_company = t.read('//*[@id="Col1-0-Profile-Proxy"]//*[contains(@class,"quote-sub-section")]/p') # read company description

        # Initialize a list to store top 5 key personnel's data (Name, Title, Pay, Year Born, Age), exclude Exercised
        key_personnel = []
        # Define the range for iterating through rows and columns
        for i in range(1, 6):  # Rows (personnel 1 to 5)
            personnel_data = []
            for j in [1,2,3,5]:  # Columns (name, title, pay, year born); range(1, 6), exclude exercised
                # Use TagUI to read the content of each cell
                cell_content = t.read('//table/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')
                personnel_data.append(cell_content)  # Append cell content to personnel_data

            # Calculate age based on year born if available
            year_born = personnel_data[-1]  # Get the year born from the last column
            if not year_born.isdigit() or len(year_born) != 4:
                age = "N/A"
            else:
                year_born = int(year_born)
                current_year = datetime.now().year
                age = current_year - year_born
            # Append age as an additional column after the year born
            personnel_data.append(age)

            key_personnel.append(personnel_data)  # Append personnel_data to key_personnel

        # Step 4: Scrape from Statistics Tab page
        t.click('//*[contains(@data-test,"STATISTICS")]') # click the statistics tab. 
        fwd_ann_div_rate = t.read('//tr[td/span[contains(text(), "Forward annual dividend rate")]]/td[2]')
        fwd_ann_div_yield = t.read('//tr[td/span[contains(text(), "Forward annual dividend yield")]]/td[2]')
        trail_ann_div_rate = t.read('//tr[td/span[contains(text(), "Trailing annual dividend rate")]]/td[2]')
        trail_ann_div_yield = t.read('//tr[td/span[contains(text(), "Trailing annual dividend yield")]]/td[2]')
        roe = t.read("//tr[contains(td/span/text(), 'Return on equity')]/td[2]")
        roa = t.read("//tr[contains(td/span/text(), 'Return on assets')]/td[2]")
        net_profit_margin = t.read("//tr[contains(td/span/text(), 'Profit margin')]/td[2]")
        p_s_ratio_ttm = t.read("//tr[contains(td/span/text(), 'Price/sales')]/td[2]")
        d_e_ratio = t.read("//tr[contains(td/span/text(), 'Total debt/equity')]/td[2]")
        current_ratio = t.read("//tr[contains(td/span/text(), 'Current ratio')]/td[2]")
        p_b_ratio = t.read("//tr[contains(td/span/text(), 'Price/book')]/td[2]")
        t.close()
    
        try:
            employee_count = int(employee_count.replace(',',''))
        except:
            employee_count = 'N/A'

        return (
            live_price, stock_market, ticker, currency, prev_close, open_price, trade_daily_volume, day_price_range, 
            day_low, day_high, market_cap, pe_ratio_ttm, eps_ttm, Company_Full, industry, employee_count, 
            fwd_ann_div_rate, fwd_ann_div_yield, trail_ann_div_rate, trail_ann_div_yield, about_company, 
            roe, roa, net_profit_margin, p_s_ratio_ttm, d_e_ratio, current_ratio, p_b_ratio, key_personnel,
            ticker_yf
        )
    
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        t.close()
        return None, None, None, None, None, None



################ 4. Create Excel workbook/ Add info to existing Excel workbook ################
def update_excel(username, company_info):
    """
    Desc: Function to create or load workbook and add data

    Input
    ------
    username (string): Username
    company_info (tuple): a tuple containing a bunch of company's finance info

    Output
    ------
    Excel workbook
    """
    # Define file path based on username
    file_path = f"{username}_stock_profile.xlsx"

    try:
        # Load existing workbook
        workbook = load_workbook(file_path)
        worksheet = workbook.active
    except FileNotFoundError:
        # If workbook doesn't exist, create a new one
        workbook = Workbook()
        worksheet = workbook.active
        # Add headers to the first row
        header = ["Date", "Live Price", "Stock Market", "Ticker Symbol", "Currency", "Previous Close", 
                  "Open Price", "Trade Daily Volume", "Day Price Range", "Day Low", "Day High", "Market Cap", 
                  "P/E Ratio TTM", "EPS TTM", "Company Name", "Industry", "Employee Count", 
                  "Fwd Annual Dividend Rate", "Fwd Annual Dividend Yield", "Trailing Annual Dividend Rate", "Trailing Annual Dividend Yield",
                  "About Company", "Return on Equity", "Return on Assets", "Net Profit Margin", "Price-to-Sales Ratio",
                  "Debt-to-equity Ratio", "Current Ratio", "Price-to-book Ratio"
                 ]
        # Add headers for key personnel
        for i in range(1, 6):
            header.extend([f"Key Personnel {i} Name", f"Key Personnel {i} Title", f"Key Personnel {i} Pay",
                           f"Key Personnel {i} Exercised", f"Key Personnel {i} Year Born", f"Key Personnel {i} Age"])
        worksheet.append(header)

    # Append data to the workbook
    # row_data = [datetime.datetime.now()] + list(company_info)
    # worksheet.append(row_data)
    # Flatten the nested list and convert elements to strings
    flatten_company_info = []
    for item in company_info:
        if isinstance(item, list):
            flatten_company_info.extend(map(str, item))
        else:
            flatten_company_info.append(str(item))
    # Append flattened data to the workbook
    # worksheet.append([datetime.datetime.now()] + flatten_company_info)
    row_data = [
        datetime.now(),
        *company_info[:28],  # First 27 elements are single-valued
    ]
    # Append key personnel data
    for personnel in company_info[28]:
        row_data.extend(personnel)

    # Extend row_data with empty strings if less than 5 key personnel are provided
#     row_data.extend([''] * (30 - len(row_data))) 

    worksheet.append(row_data)
    

    # Save the workbook
    workbook.save(file_path)




################ 5. Scrape for Financial Info via Google (Based on Generic Benchmark; Rules) ################
def convert_to_float(value):
    if "%" in value:
        return float(value.strip("%")) / 100
    return float(value)

def get_performance(ratio_value, ratio_name):
    value = convert_to_float(ratio_value)
    if ratio_name == "ROE":
        if value < 0.05:
            return "Poor"
        elif 0.05 <= value < 0.1:
            return "Fair"
        elif 0.1 <= value < 0.2:
            return "Good"
        else:
            return "Excellent"
    elif ratio_name == "ROA":
        if value < 0.02:
            return "Poor"
        elif 0.02 <= value < 0.05:
            return "Fair"
        elif 0.05 <= value < 0.1:
            return "Good"
        else:
            return "Excellent"
    elif ratio_name == "Net Profit Margin":
        if value < 0.05:
            return "Poor"
        elif 0.05 <= value < 0.1:
            return "Fair"
        elif 0.1 <= value < 0.2:
            return "Good"
        else:
            return "Excellent"
    elif ratio_name == "Price to Sales Ratio":
        if value > 10:
            return "Poor"
        elif 5 <= value <= 10:
            return "Fair"
        elif 2 <= value < 5:
            return "Good"
        else:
            return "Excellent"
    elif ratio_name == "Debt to Equity Ratio":
        if value > 2:
            return "Poor"
        elif 1 <= value <= 2:
            return "Fair"
        elif 0.5 <= value < 1:
            return "Good"
        else:
            return "Excellent"
    elif ratio_name == "Current Ratio":
        if value < 1:
            return "Poor"
        elif 1 <= value <= 1.5:
            return "Fair"
        elif 1.5 <= value <= 2:
            return "Good"
        else:
            return "Excellent"
    elif ratio_name == "Price to Book Ratio":
        if value > 4:
            return "Poor"
        elif 1 <= value <= 4:
            return "Fair"
        elif 0.5 <= value < 1:
            return "Good"
        else:
            return "Excellent"
    elif ratio_name == "Price to Earnings Ratio":
        if value > 30:
            return "Poor"
        elif 15 <= value <= 30:
            return "Fair"
        elif 10 <= value < 15:
            return "Good"
        else:
            return "Excellent"
    else:
        return "Unknown"


def performance_analysis(financial_ratios, buy_threshold=18, sell_threshold=12):
    """
    Desc: To analysis company financial's performance based on buy/ sell threshold

    Input
    ------
    company_info (list): A bunch of company's information
    buy_threshold: threshold that advices users to buy stocks if it crosses
    sell_threshold: threshold that advices users to sell stocks if it goes below

    Returns
    ------
    ratios_results (dict): Abunch of analyses over company's financial performance
    """

    # financial_ratios = [company_info[21], company_info[22], company_info[23], company_info[24], company_info[25], company_info[26], company_info[27], company_info[11]]

    ratios = ["ROE", "ROA", "Net Profit Margin", "Price to Sales Ratio", "Debt to Equity Ratio",
            "Current Ratio", "Price to Book Ratio", "Price to Earnings Ratio"]
    
    # print(f"financial_ratios:{financial_ratios}")
    # print(f'financial_ratios type:{type(financial_ratios)}')

    ratios_results = {}
    performance_result = []

    for i in range(len(ratios)):
        performance = get_performance(financial_ratios[i], ratios[i])
        performance_scores = {"Excellent": 3, "Good": 2, "Fair": 1, "Poor": 0}
        total_score = sum(performance_scores[get_performance(financial_ratios[i], ratios[i])] for i in range(len(ratios)))
        ratios_results[f'{ratios[i]} Performance'] = performance

        # print(f"{ratios[i]} Performance: {performance}")

        performance_result.append(performance)
        
    # Determine recommendation
    if total_score >= buy_threshold:
        recommendation = "Buy"
    elif sell_threshold <= total_score < buy_threshold:
        recommendation = "Hold"
    else:
        recommendation = "Sell"

    # print("Overall Recommendation:", recommendation)
    ratios_results['Overall Recommendation'] = recommendation

    ratios.append('Overall Recommendation')
    performance_result.append(recommendation)

    ratios_results = {'Ratio': ratios, 'Performance': performance_result}

    return ratios_results


def scrape_n_return(company):
    #Scrape for financial info about company via Yahoo

    # username = "ky"  # Define username
    # company = 'apple' # User input company of interest
    # chart_duration = ["1d", "1y"] # set as default. Options of "1d", "5d", "1m", "6m", "YTD", "1y", "5y", and "Max".
    try:
        company_info = extract_company_info_yfinance(company) # chart_duration; stores extracted information
    except:
        print(f'Something happened when scrapping')
    if company_info[0]:  # Check if data is not None
        (
            live_price, stock_market, ticker, currency, prev_close, open_price, trade_daily_volume, day_price_range, 
            day_low, day_high, market_cap, pe_ratio_ttm, eps_ttm, Company_Full, industry, employee_count, 
            fwd_ann_div_rate, fwd_ann_div_yield, trail_ann_div_rate, trail_ann_div_yield, about_company, 
            roe, roa, net_profit_margin, p_s_ratio_ttm, d_e_ratio, current_ratio, p_b_ratio, key_personnel,
            ticker_yf
        ) = company_info
        # Process retrieved data
        print('Company information retrieval succeeded!')
        return company_info
    else:
        print("Company information retrieval failed.")
        return None

    # update_excel(username, company_info)



################ 6. Scrape for News/ Get News via API ################
def news_scrape_API(target, when, howmany=8):
    # To scrape for news using API key
    # api_key = 'aff693b1fa0146caa57ac92ed9173386'
    # api_key = '5a06a0e119fa4dbda94d57f298b1b5a1'
    api_key = os.environ.get("NEWS_API_KEY")
    url = f'https://newsapi.org/v2/everything?q={target}&apiKey={api_key}&pageSize={howmany}&from={when}&sortBy=publishedAt&language=en'

    # Initialize an empty list to store the titles and descriptions
    # output = []
    titles = []
    descriptions = []
    all_articles = []

    try:
        # Make the API request
        response = requests.get(url)
        data = response.json()

        # Check if the request was successful
        if response.status_code == 200:
            # Extract title and description for each article
            articles = data['articles']
            for article in articles:
                title = article['title']
                description = article['description']
                # output.append({'title': title, 'description': description})
                titles.append(title)
                descriptions.append(description)
                all_articles.append(title+'. '+ description)
        else:
            print("Error fetching data from the API:", data.get('message', ''))
    except Exception as e:
        print("An error occurred during the API request:", e)

    return titles, descriptions, all_articles


def news_scrape_rpa(start_date_str,end_date_str,company='Apple',num_of_articles=8, num_of_page=1):
    # Web automation to scrape for news on Google

    t.init()
    t.url('https://www.google.com/webhp?hl=en&sa=X&ved=0ahUKEwj9ipr4jaCFAxWSSGcHHdDhCHoQPAgI')
    t.type('//*[@id="APjFqb"]', company) #type company
    t.wait(2)
    # t.click('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[2]/div[4]/div[7]/center/input[1]') #click search
    t.click('//center/input[@role="button"]') #click search
    # t.click('//*[@id="hdtb-sc"]/div/div/div[1]/div[4]/a/div') #click news
    t.click('//div[contains(text(), "News")]')
    t.wait(1)
    # t.click('//*[@id="hdtb-tls"]') # Tools
    t.click('//div[contains(text(), "Tools")]')
    t.wait(1)
    t.click('//*[@id="tn_1"]/span[1]/g-popup/div[1]/div/div/div') #Click on time setting
    t.wait(1)
    # t.click('//*[@id="lb"]/div/g-menu/g-menu-item[3]/div/a') #Click on past 24 hours
    t.click('//*[@id="lb"]/div/g-menu/g-menu-item[8]') #Click on Custom range ; '//*[@id="lb"]/div/g-menu/g-menu-item[3]/div/a'
    t.wait(1)
    t.type('//*[@id="OouJcb"]',start_date_str) # Type 'From' date
    t.type('//*[@id="rzG2be"]',end_date_str) # Type 'To' date
    t.wait(1)
    t.click('//*[@id="T3kYXe"]/g-button')

    # Initialize an empty list to store all articles
    all_articles = []
    titles = []
    summaries = []
    dates = []

    for page in range(1, num_of_page + 1):
        t.wait(2)
        titles_xpath = '//div[@aria-level="3" and @role="heading"]'  # XPaths
        summaries_xpath = '//div[contains(@class, "GI74Re") and contains(@class, "nDgy9d")]'
        dates_xpath = '//div[@style="bottom:0px"]'

        # Fetch titles and summaries
        for i in range(1, num_of_articles + 1):
            title_xpath = '(' + titles_xpath + ')[' + str(i) + ']'
            titles.append(t.read(title_xpath))
            summary_xpath = '(' + summaries_xpath + ')[' + str(i) + ']'
            summaries.append(t.read(summary_xpath).replace('\n',''))
            date_xpath = '(' + dates_xpath + ')['+ str(i) + ']'
            dates.append(t.read(date_xpath))

        # Store results for this page in the all_articles list
        for i in range(len(titles)):
            summary_clean = summaries[i].replace('\n', ' ')
            article = titles[i] + ". " + summary_clean
            all_articles.append(article)

        # Click the "Next" button to go to the next page, if not on the last page
        if page < num_of_page:
            next_button_xpath = '//*[@id="pnnext"]'  # XPath for "Next" button
            if t.present(next_button_xpath):
                t.click(next_button_xpath)
            else:
                print("No more pages to navigate.")
                break

    # Close TagUI
    t.close()

    return titles, summaries, dates, all_articles


def similarity(all_articles, threshold):
    # Initialize the TF-IDF Vectorizer
    tfidf_vectorizer = TfidfVectorizer()

    for _ in range(2):  # Run the duplicate removal process twice
        # Fit and transform the articles to TF-IDF vectors
        tfidf_matrix = tfidf_vectorizer.fit_transform(all_articles)

        # List to keep track of articles to keep
        keep = set(range(len(all_articles)))

        # Compare similarity between all pairs of articles
        for i, j in combinations(range(len(all_articles)), 2):
            # Calculate cosine similarity between the two TF-IDF vectors
            similarity_score = cosine_similarity(tfidf_matrix[i], tfidf_matrix[j])[0][0]

            # Check if similarity score exceeds the threshold
            if similarity_score > threshold:
                # Prefer to remove the latter article in the pair
                keep.discard(j)

        # Create a new list with only the unique articles
        all_articles = [all_articles[i] for i in sorted(keep)]

    return all_articles

# Define a function to determine sentiment label based on compound score
def get_sentiment_label(sentiment_scores):
    compound_score = sentiment_scores['compound']
    if compound_score >= 0.52:
        return 'Positive'
    elif compound_score <= -0.48:
        return 'Negative'
    else:
        return 'Neutral'

# Assume all_articles is available from the previous step
def sentiment(all_articles):
    # Initialize VADER sentiment analyzer
    sid = SentimentIntensityAnalyzer()

    # Variables to store counts of positive, negative, and neutral sentiments
    positive_count = 0
    negative_count = 0
    neutral_count = 0
    total_sentiment_score = 0

    # Perform sentiment analysis on each article
    for article in all_articles:
        # Get sentiment scores for the article
        sentiment_scores = sid.polarity_scores(article)
        sentiment_label = get_sentiment_label(sentiment_scores)

        # Update counts based on sentiment label
        if sentiment_label == 'Positive':
            positive_count += 1
            total_sentiment_score += 1
        elif sentiment_label == 'Negative':
            negative_count += 1
            total_sentiment_score -= 1
        else:
            neutral_count += 1

    # Calculate average sentiment score
    total_articles = positive_count + negative_count + neutral_count
    average_sentiment_score = (positive_count - negative_count) / total_articles if total_articles > 0 else 0
    sent = average_sentiment_score
    positive = int(positive_count)
    negative = int(negative_count)
    neutral = int(neutral_count)
    return sent, positive, negative, neutral


def interpret_sentiment_v2(API_articles, RPA_articles, threshold): # total_sentiment, total_positive, total_negative, total_neutral
    
    try:
        unique_articles_API = similarity(API_articles, threshold)
        API_sent, API_positive, API_negative, API_neutral = sentiment(unique_articles_API)
    except:
        API_sent = 0
        API_positive = 0
        API_negative = 0
        API_neutral = 0
    unique_articles_RPA = similarity(RPA_articles, threshold)
    RPA_sent, RPA_positive, RPA_negative, RPA_neutral = sentiment(unique_articles_RPA)

    # Total_articles = RPA_positive + RPA_negative + RPA_neutral
    # Total_sentiment = (RPA_positive - RPA_negative) / Total_articles if Total_articles > 0 else 0

    Total_positive = API_positive+RPA_positive
    Total_negative = API_negative+RPA_negative
    Total_neutral = API_neutral+RPA_neutral

    Total_articles = Total_positive+Total_negative+Total_neutral

    Total_API = API_positive+API_negative+API_neutral
    Total_RPA = RPA_positive+RPA_negative+RPA_neutral
    Total_sentiment = (Total_positive - Total_negative) / Total_articles if Total_articles > 0 else 0


    # Interpret the combined average sentiment score
    if Total_sentiment > 0.02:
        sentiment_summary = "Overall Positive Sentiment"
    elif Total_sentiment < -0.02:
        sentiment_summary = "Overall Negative Sentiment"
    else:
        sentiment_summary = "Neutral Sentiment"

    summary = f"\
    Overall Sentiment Report\n\
    ---\n\
    No. of API articles: {int(Total_API)}\n\
    No. of RPA articles: {int(Total_RPA)}\n\
    Total No. of articles: {int(Total_articles)}\n\
    \n\
    No. of Positive articles: {Total_positive}\n\
    No. of Negative articles: {Total_negative}\n\
    No. of Neutral articles: {Total_neutral}\n\
    \n\
    API Sentimate: {API_sent}\n\
    RPA Sentimate : {RPA_sent}\n\
    Combined Average Sentiment Score: {Total_sentiment}\n\
    "

    # return summary
    return [Total_positive, Total_negative, Total_neutral, Total_articles, Total_sentiment]




# def interpret_sentiment(all_articles, threshold): # total_sentiment, total_positive, total_negative, total_neutral
    
#     unique_articles_RPA = similarity(all_articles, threshold)
#     RPA_sent, RPA_positive, RPA_negative, RPA_neutral = sentiment(unique_articles_RPA)

#     Total_articles = RPA_positive + RPA_negative + RPA_neutral

#     Total_sentiment = (RPA_positive - RPA_negative) / Total_articles if Total_articles > 0 else 0

#     # Interpret the combined average sentiment score
#     if Total_sentiment > 0.02:
#         sentiment_summary = "Overall Positive Sentiment"
#     elif Total_sentiment < -0.02:
#         sentiment_summary = "Overall Negative Sentiment"
#     else:
#         sentiment_summary = "Neutral Sentiment"

#     summary = f"\
#     Overall Sentiment Report\n\
#     ---\n\
#     Total Number of Positive-Sentiment Articles: {RPA_positive} \n\
#     Total Number of Negative-Sentiment Articles: {RPA_negative} \n\
#     Total Number of Neutral-Sentiment Articles: {RPA_neutral} \n\
#     \n\
#     Total Number of Articles: {Total_articles} \n\
#     \n\
#     Sentiment Interpretation: {Total_sentiment}\
#     "

#     # return summary
#     return [RPA_positive, RPA_negative, RPA_neutral, Total_articles, Total_sentiment]


if __name__ == "__main__":

    t.close()


# ######### 1. Comparing scraping results between TagUI and Yfinance python package. 
# # This will be a means to validate results from both TagUI and yfinance python package which is not within our control
# # Note that Minor Differences will occur when the stocks are being traded
# company = 'singtel' # User input company of interest
# tagui_yfinance_validation(company)


# ######### 2. Stock Price Chart Snapshot 
# chart_duration = ["1d", "1y"] # Options of "1d", "5d", "1m", "6m", "YTD", "1y", "5y", and "Max". Default set to 1 day and 1 year charts
# chart_snapshot(chart_duration)


# ######### 3. Extract company info
# username = "ky"  # Define username
# company = 'apple' # User input company of interest
# chart_duration = ["1d", "1y"] # set as default. Options of "1d", "5d", "1m", "6m", "YTD", "1y", "5y", and "Max".

# company_info = extract_company_info_yfinance(company, chart_duration) # stores extracted information
# if company_info[0]:  # Check if data is not None
#     (
#         live_price, stock_market, ticker, currency, prev_close, open_price, trade_daily_volume, day_price_range, 
#         day_low, day_high, market_cap, pe_ratio_ttm, eps_ttm, Company_Full, industry, employee_count, 
#         fwd_ann_div_rate, fwd_ann_div_yield, trail_ann_div_rate, trail_ann_div_yield, about_company, 
#         roe, roa, net_profit_margin, p_s_ratio_ttm, d_e_ratio, current_ratio, p_b_ratio, key_personnel
#     ) = company_info
#     # Process retrieved data
# else:
#     print("Company information retrieval failed.")

# update_excel(username, company_info)
    

# ######### 3.1 Technical Analysis
# performance_analysis(company_info, buy_threshold=18, sell_threshold=12)
